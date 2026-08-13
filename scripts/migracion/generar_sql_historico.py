"""
Genera el SQL de migración de los 92 registros reales de
docs/prog_cx.xlsx hacia la tabla solicitudes_cirugia.

Uso: uv run --with openpyxl python generar_sql_historico.py
Escribe 0006_datos_historicos.sql (UTF-8) en esta misma carpeta.
"""
import re
import unicodedata
from datetime import datetime
import openpyxl


def sin_acentos(s):
    return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))

XLSX = "../../docs/prog_cx.xlsx"
SALIDA = "0006_datos_historicos.sql"

ESPECIALIDAD_POR_HOJA = {
    "ORTOPEDIA": "Ortopedia",
    "NEUROCIRUGIA": "Neurocirugía",
    "CIRUGIA MAXILOFACIAL": "Cirugía Maxilofacial",
    "CARDIOVASCULAR": "Cardiovascular",
    "UROLOGIA": "Urología",
    "GINECOLOGIA": "Ginecología",
    "CIRUGIA GENERAL": "Cirugía General",
}

EPS_CANONICAS = {
    "ADRES": "ADRES", "ASMET SALUD": "ASMET SALUD", "EMSSANAR": "EMSSANAR",
    "NEPS": "NEPS", "NUEVA EPS": "NUEVA EPS", "SANIDA MILITAR": "SANIDAD MILITAR",
    "SANIDAD MILITAR": "SANIDAD MILITAR", "SOAT": "SOAT", "SOS": "SOS", "SURA": "SURA",
}

# Prefijos de unidad evaluados en orden (UCIN antes que UCI para no truncar)
UNIDAD_PREFIJOS = [
    ("UCIN", "UCIN"),
    ("UCI", "UCI"),
    ("URG", "Urgencias"),
    ("HOSP", "Hospitalización"),
    ("AMBULATORIO", "Ambulatorio"),
    ("RECUPERACION", "Recuperación"),
    ("RECUPERACIÓN", "Recuperación"),
]

REMANENTE_IRRELEVANTE = {"CIRUGIA", "CIRUGÍA", "QUIRURGICA", "QUIRÚRGICA", "ADULTO"}


def sql_str(v):
    if v is None:
        return "null"
    s = str(v).strip()
    if s == "":
        return "null"
    return "'" + s.replace("'", "''") + "'"


def sql_int(v):
    if v is None:
        return "null"
    try:
        return str(int(v))
    except (ValueError, TypeError):
        return "null"


def sql_bool(v):
    return "true" if v else "false"


def limpiar_eps(raw):
    if not raw:
        return None
    key = re.sub(r"\s+", " ", str(raw).strip().upper())
    key = key.split(" // ")[0].strip()  # variante rara "SOAT // EMSSANAR": se queda con la primera
    return EPS_CANONICAS.get(key)


def separar_ubicacion(raw):
    if not raw:
        return None, None
    s = re.sub(r"\s+", " ", str(raw).strip())
    upper = s.upper()
    tokens = upper.split(" ", 1)
    token0 = tokens[0]
    resto = tokens[1].strip() if len(tokens) > 1 else ""

    for prefijo, unidad in UNIDAD_PREFIJOS:
        if token0.startswith(prefijo):
            if resto in REMANENTE_IRRELEVANTE or resto == "":
                return unidad, None
            return unidad, resto

    # Sin coincidencia de unidad conocida: se asume número de cama en Hospitalización
    return "Hospitalización", s


def detectar_estado(*campos_texto):
    texto = " ".join(c for c in campos_texto if c).upper()
    if "CANCELA" in texto:
        return "cancelado", texto
    if "REALIZAD" in texto:
        return "realizado", None
    if "PROGRAMADO" in texto or "PROGRAMAR" in texto:
        return "programado", None
    return "procesado", None


def header_row(ws):
    for r in (1, 2):
        v = ws.cell(row=r, column=1).value
        if v and isinstance(v, str) and "fecha" in v.lower():
            return r
    return 1


def procesar_hoja(ws, nombre_hoja):
    filas_sql = []
    hr = header_row(ws)
    headers = {}
    subespecialidad_actual = None

    for r in range(1, ws.max_row + 1):
        if r <= hr:
            if r == hr:
                for c in range(1, ws.max_column + 1):
                    h = ws.cell(row=r, column=c).value
                    if h:
                        # Cirugía General repite el bloque de encabezados en las
                        # columnas L:V (100% vacío) — se conserva solo la 1ra
                        # ocurrencia (bloque real A:K) al construir el mapeo.
                        key = sin_acentos(re.sub(r"\s+", " ", str(h).strip().lower()))
                        if key not in headers:
                            headers[key] = c
            continue

        col_doc = ws.cell(row=r, column=headers.get("documento", 2)).value

        # Encabezado repetido en medio de la hoja (p.ej. Ortopedia fila 28)
        if isinstance(col_doc, str) and col_doc.strip().lower() == "documento":
            continue

        col1 = ws.cell(row=r, column=1).value
        col2 = ws.cell(row=r, column=2).value
        # Sub-encabezado tipo "ORTOPEDIA RECONSTRUCTIVA" (fila con solo texto en col1)
        if isinstance(col1, str) and col2 is None and col1.strip() and "fecha" not in col1.lower():
            texto = col1.strip()
            if texto.isupper() and len(texto) > 4:
                subespecialidad_actual = texto.title()
                continue

        documento = col_doc
        if documento in (None, ""):
            continue

        def val(*keys):
            for k in keys:
                c = headers.get(sin_acentos(k))
                if c:
                    v = ws.cell(row=r, column=c).value
                    if v not in (None, ""):
                        return v
            return None

        fecha_reporte = val("fecha", "fecha de solicitud del turno")
        numero_ingreso = val("#ingreso")
        nombre = val("nombre")
        edad = val("edad")
        eps_raw = val("eps")
        ubicacion_raw = val("ubicación")
        preanestesica = val("cuenta con valoración preanestesica")
        procedimiento = val("procedimiento", "procedimiento quirurgico")
        autorizacion = val("autorizacion de aseguradora")
        tiempo_est = val("tiempo quirúrgico estimado")
        especialista = val("especialista")
        fecha_cirugia_txt = val("fecha de cirugía", "fecha de cirgia")
        obs_prog = val("observaciones programación")
        obs = val("observaciones")
        obs_orserv = val("orservaciones")
        estado_material = val("estado material de osteosintesis")
        casa_medica = val("casa medica que entrega material")

        unidad, cama = separar_ubicacion(ubicacion_raw)
        eps_nombre = limpiar_eps(eps_raw)

        piezas_obs = []
        if fecha_cirugia_txt and str(fecha_cirugia_txt).strip():
            piezas_obs.append(str(fecha_cirugia_txt).strip())
        if obs_prog:
            piezas_obs.append(str(obs_prog).strip())
        if obs:
            piezas_obs.append(str(obs).strip())
        if obs_orserv:
            piezas_obs.append(str(obs_orserv).strip())
        if especialista:
            piezas_obs.append(f"Especialista: {str(especialista).strip()}")
        observaciones_final = " | ".join(p for p in piezas_obs if p) or None

        estado, motivo_cancelacion = detectar_estado(*piezas_obs)

        if not numero_ingreso:
            numero_ingreso = f"HIST-{nombre_hoja[:4]}-{r}"

        fecha_reporte_sql = "now()"
        if isinstance(fecha_reporte, datetime):
            fecha_reporte_sql = sql_str(fecha_reporte.strftime("%Y-%m-%d 12:00:00-05"))

        especialidad_nombre = ESPECIALIDAD_POR_HOJA[nombre_hoja]
        subesp_sql = sql_str(subespecialidad_actual) if nombre_hoja == "ORTOPEDIA" else "null"

        eps_sub = f"(select id from eps where nombre = {sql_str(eps_nombre)})" if eps_nombre else "null"
        unidad_sub = f"(select id from unidades where nombre = {sql_str(unidad)})" if unidad else "null"

        fila = f"""(
  {sql_str(numero_ingreso)}, {sql_str(documento)},
  (select id from especialidades where nombre = {sql_str(especialidad_nombre)}), {subesp_sql},
  (select id from perfiles where email = 'juan.etayo@cacsantabarbara.co'), {fecha_reporte_sql},
  {sql_str(estado)},
  {sql_str(nombre)}, {sql_int(edad)}, {eps_sub}, {unidad_sub}, {sql_str(cama)},
  {sql_str(preanestesica)}, {sql_str(procedimiento)}, {sql_int(tiempo_est)},
  {sql_str(autorizacion)}, {sql_str(observaciones_final)}, {sql_str(estado_material)}, {sql_str(casa_medica)},
  {sql_bool(motivo_cancelacion is not None)}, {sql_str(motivo_cancelacion)},
  true
)"""
        filas_sql.append(fila)
    return filas_sql


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    columnas = """(
  numero_ingreso, documento_paciente,
  especialidad_id, subespecialidad,
  reportado_por, fecha_reporte,
  estado,
  nombre_paciente, edad, eps_id, unidad_id, cama,
  valoracion_preanestesica, procedimiento, tiempo_estimado_minutos,
  autorizacion_aseguradora, observaciones_programacion, estado_material_osteosintesis, casa_medica_material,
  cancelado, motivo_cancelacion,
  es_historico
)"""
    todas = []
    for hoja in ESPECIALIDAD_POR_HOJA:
        ws = wb[hoja]
        todas.extend(procesar_hoja(ws, hoja))

    with open(SALIDA, "w", encoding="utf-8") as f:
        f.write("-- Migración histórica generada automáticamente desde docs/prog_cx.xlsx\n")
        f.write(f"-- Total de registros: {len(todas)}\n")
        f.write(f"insert into solicitudes_cirugia {columnas} values\n")
        f.write(",\n".join(todas) + ";\n")

    print(f"Generados {len(todas)} registros en {SALIDA}")


if __name__ == "__main__":
    main()
